"""
DEPRECATED: This module is deprecated and will be removed in a future version.

Use the following services instead:
- app.services.pdf_generator.PDFGenerator for PDF generation
- app.services.excel_generator.ExcelGenerator for Excel generation

These services return BytesIO buffers (more secure) instead of writing files to disk.
"""
import warnings
import os
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from typing import List, Dict, Any
import datetime
from app.core.config import settings
import logging

logger = logging.getLogger(__name__)

# SECURITY WARNING: Writing files to disk can be a security risk.
# Consider using PDFGenerator and ExcelGenerator which return BytesIO buffers instead.
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "../../../data/exports")
os.makedirs(OUTPUT_DIR, exist_ok=True)

def _emit_deprecation_warning(func_name: str):
    """Emit deprecation warning for legacy functions."""
    msg = (
        f"{func_name}() is deprecated and will be removed in a future version. "
        f"Use app.services.pdf_generator.PDFGenerator or "
        f"app.services.excel_generator.ExcelGenerator instead."
    )
    warnings.warn(msg, DeprecationWarning, stacklevel=3)
    logger.warning(msg)

def generate_pdf_quotation(quotation_id: str, items_data: Any, total_cost: float) -> str:
    """
    DEPRECATED: Use app.services.pdf_generator.PDFGenerator instead.

    Generates a professional PDF quotation with detailed material descriptions and implementation details.
    """
    _emit_deprecation_warning("generate_pdf_quotation")
    filename = f"quotation_{quotation_id}_{datetime.datetime.now().strftime('%Y%m%d')}.pdf"
    filepath = os.path.join(OUTPUT_DIR, filename)
    
    # Flatten items if it's a dict (materials/labor)
    items = []
    if isinstance(items_data, dict):
        if "materials" in items_data:
            items.extend(items_data["materials"].get("items", []))
        if "labor" in items_data:
            items.extend(items_data["labor"].get("trades", []))
    elif isinstance(items_data, list):
        items = items_data
    
    c = canvas.Canvas(filepath, pagesize=A4)
    width, height = A4
    
    # Header
    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, height - 50, f"Construction Quotation #{quotation_id}")
    
    c.setFont("Helvetica", 10)
    c.drawString(50, height - 70, f"Date: {datetime.date.today().strftime('%B %d, %Y')}")
    c.drawString(50, height - 85, "Status: Professional Estimate")
    
    # Table Headers with separate columns for name and description
    y = height - 120
    c.setFont("Helvetica-Bold", 10)
    c.drawString(30, y, "No")
    c.drawString(50, y, "Material/Item")
    c.drawString(200, y, "Description & Implementation")
    c.drawString(400, y, "Unit")
    c.drawString(450, y, "Qty")
    c.drawString(500, y, "Unit Price")
    c.drawString(550, y, "Total (EGP)")
    
    c.line(30, y-5, 570, y-5)
    
    y -= 25
    c.setFont("Helvetica", 8)
    
    for i, item in enumerate(items, 1):
        if not isinstance(item, dict): continue
        
        # Get material name and description separately
        material_name = item.get("name") or item.get("trade", "Unknown Item")
        description = item.get("description") or material_name
        
        # Truncate material name if too long
        from app.core.config import settings
        max_name_len = settings.MAX_MATERIAL_NAME_LENGTH
        if len(material_name) > max_name_len:
            material_name = material_name[:max_name_len - 3] + "..."
        
        # Wrap description into multiple lines (max 3 lines)
        desc_lines = []
        max_desc_width = settings.MAX_DESCRIPTION_WIDTH  # characters per line
        words = description.split()
        current_line = ""
        
        for word in words:
            if len(current_line + " " + word) <= max_desc_width:
                current_line += (" " + word if current_line else word)
            else:
                if current_line:
                    desc_lines.append(current_line)
                current_line = word
        if current_line:
            desc_lines.append(current_line)
        
        # Limit to 3 lines max, truncate last line if needed
        desc_lines = desc_lines[:3]
        if len(desc_lines) == 3 and len(desc_lines[2]) > max_desc_width - 3:
            desc_lines[2] = desc_lines[2][:max_desc_width-3] + "..."
        
        unit = item.get("unit", "ea")
        qty = item.get("quantity") or item.get("hours", 0)
        price = item.get("unit_price") or item.get("rate", 0)
        line_total = item.get("total") or (float(qty) * float(price))
        
        # Draw item number
        c.drawString(30, y, str(i))
        
        # Draw material name
        c.drawString(50, y, material_name)
        
        # Draw description (multi-line)
        desc_y = y
        for line in desc_lines:
            c.drawString(200, desc_y, line)
            desc_y -= 12
        
        # Draw unit, quantity, price, total
        c.drawString(400, y, str(unit))
        c.drawString(450, y, f"{qty:.2f}")
        c.drawString(500, y, f"{price:.2f}")
        c.drawString(550, y, f"{line_total:,.2f}")
        
        # Move to next item (account for multi-line description)
        y -= max(20, len(desc_lines) * 12 + 5)
        
        # New page if needed
        if y < 100:
            c.showPage()
            y = height - 50
            c.setFont("Helvetica", 8)
            
    # Total line
    c.line(30, y, 570, y)
    y -= 25
    c.setFont("Helvetica-Bold", 11)
    c.drawString(400, y, "Grand Total:")
    c.drawString(550, y, f"{total_cost:,.2f} EGP")
    
    c.save()
    return filepath

def generate_excel_quotation(quotation_id: str, items_data: Any, total_cost: float) -> str:
    """
    DEPRECATED: Use app.services.excel_generator.ExcelGenerator instead.

    Generates a professional Excel quotation with detailed material descriptions and implementation details.
    """
    _emit_deprecation_warning("generate_excel_quotation")
    filename = f"quotation_{quotation_id}_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    
    # Flatten items if it's a dict
    items = []
    if isinstance(items_data, dict):
        if "materials" in items_data:
            items.extend(items_data["materials"].get("items", []))
        if "labor" in items_data:
            items.extend(items_data["labor"].get("trades", []))
    elif isinstance(items_data, list):
        items = items_data
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Professional BOQ"
    
    # Style headers
    from openpyxl.styles import Font, Alignment, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    ws.append(["Quotation ID", quotation_id])
    ws.append(["Date", datetime.date.today()])
    ws.append([])
    
    # 6-Column Header with enhanced description column
    headers = ["Item No", "Description & Implementation (الوصف)", "Unit (وحدة)", "Unit Price (EGP)", "Quantity (MOQ)", "Total (EGP)"]
    ws.append(headers)
    
    for cell in ws[4]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    for i, item in enumerate(items, 1):
        if not isinstance(item, dict): continue
        
        # Get material name and description - combine them for comprehensive description
        material_name = item.get("name") or item.get("trade", "Unknown")
        description = item.get("description") or material_name
        
        # If description doesn't include material name, prepend it
        if material_name and material_name not in description:
            full_description = f"{material_name}: {description}"
        else:
            full_description = description
        
        unit = item.get("unit") or ("hr" if "trade" in item else "-")
        qty = item.get("quantity") or item.get("hours", 0)
        price = item.get("unit_price") or item.get("rate", 0)
        total = item.get("total") or (float(qty) * float(price))
        
        row = [i, full_description, unit, price, qty, total]
        ws.append(row)
        
        # Wrap text for description column (column B) with proper alignment
        ws.cell(row=ws.max_row, column=2).alignment = Alignment(wrap_text=True, horizontal="right", vertical="top")

    ws.append([])
    ws.append(["", "", "", "", "Grand Total", total_cost])
    ws[ws.max_row][5].font = Font(bold=True)
    ws[ws.max_row][6].font = Font(bold=True)

    # Adjust column widths for better readability
    ws.column_dimensions['A'].width = 10  # Item No
    ws.column_dimensions['B'].width = 80  # Description & Implementation
    ws.column_dimensions['C'].width = 15  # Unit
    ws.column_dimensions['D'].width = 15  # Unit Price
    ws.column_dimensions['E'].width = 15  # Quantity
    ws.column_dimensions['F'].width = 20  # Total
    
    # Set row height for description rows to accommodate wrapped text
    for row_idx in range(5, ws.max_row):
        ws.row_dimensions[row_idx].height = 60  # Increased height for wrapped descriptions
    
    wb.save(filepath)
    return filepath
