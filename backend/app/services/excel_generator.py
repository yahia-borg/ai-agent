"""
Excel export service for quotations
Generates multi-sheet Excel files with detailed BOQ-style breakdowns.
Hierarchical BOQ format grouped by trade/category with bilingual headers.
"""
from io import BytesIO
from typing import Dict, Any, Optional, List, Tuple
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from app.models.quotation import Quotation, QuotationData
from app.utils.language_detector import detect_language


# ── Bilingual category names for header rows ────────────────────────
_CATEGORY_NAMES_BILINGUAL: Dict[str, Dict[str, str]] = {
    "flooring": {"en": "Flooring Works", "ar": "اعمال الارضيات"},
    "painting": {"en": "Painting Works", "ar": "اعمال الدهانات"},
    "plastering": {"en": "Plastering Works", "ar": "اعمال البياض"},
    "plumbing": {"en": "Plumbing Works", "ar": "اعمال السباكة"},
    "electrical": {"en": "Electrical Works", "ar": "اعمال الكهرباء"},
    "carpentry": {"en": "Carpentry & Joinery", "ar": "اعمال النجارة"},
    "demolition": {"en": "Demolition Works", "ar": "اعمال الهدم"},
    "labor": {"en": "Labor", "ar": "العمالة"},
    "permits": {"en": "Permits & Fees", "ar": "التراخيص والرسوم"},
    "general": {"en": "General Works", "ar": "اعمال عامة"},
    "materials": {"en": "Materials", "ar": "المواد"},
}


class ExcelGenerator:
    """Generate Excel quotation documents with detailed BOQ format"""

    def __init__(self):
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.category_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.category_font = Font(bold=True, color="FFFFFF", size=11)
        self.row_even_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        self.title_font = Font(bold=True, size=16)
        self.subtitle_font = Font(bold=True, size=12)
        self.bold_font = Font(bold=True, size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.thick_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
    
    def _normalize_cost_breakdown(self, cost_breakdown: Any, total_cost: float) -> Dict[str, Any]:
        """
        Normalize cost_breakdown to expected dictionary format.
        Handles both list format (from create_quotation tool) and dictionary format.
        """
        if isinstance(cost_breakdown, list):
            # Convert list of items to structured dictionary format
            items_total = sum(item.get("quantity", 0) * item.get("unit_price", 0) for item in cost_breakdown if isinstance(item, dict))
            
            return {
                "currency": "EGP",
                "materials": {
                    "subtotal": items_total,
                    "percentage": 100.0 if total_cost > 0 else 0.0,
                    "items": [
                        {
                            "name": item.get("name", ""),
                            "quantity": item.get("quantity", 0),
                            "unit_cost": item.get("unit_price", 0),
                            "cost": item.get("quantity", 0) * item.get("unit_price", 0),
                            "unit": item.get("unit", ""),
                            "description": item.get("description", item.get("name", "")),
                            "category": item.get("category", "General")
                        }
                        for item in cost_breakdown if isinstance(item, dict)
                    ]
                }
            }
        elif isinstance(cost_breakdown, dict):
            # Already in dictionary format, return as-is
            return cost_breakdown
        else:
            # Fallback: empty structure
            return {"currency": "EGP"}
    
    @staticmethod
    def _detect_item_category(item: Dict[str, Any]) -> str:
        """Detect the trade/category key for grouping an item."""
        cat = str(item.get("category", "")).lower()
        name = str(item.get("name", "")).lower()
        combined = f"{cat} {name}"
        for key in ["flooring", "painting", "plastering", "plumbing", "electrical", "carpentry", "demolition"]:
            if key in combined:
                return key
        if any(k in combined for k in ["tile", "ceramic", "porcelain", "marble", "parquet"]):
            return "flooring"
        if any(k in combined for k in ["paint", "دهان"]):
            return "painting"
        if any(k in combined for k in ["plaster", "محارة", "بياض"]):
            return "plastering"
        if any(k in combined for k in ["pipe", "toilet", "sink", "سباكة"]):
            return "plumbing"
        if any(k in combined for k in ["cable", "wire", "switch", "كهرباء"]):
            return "electrical"
        if any(k in combined for k in ["door", "window", "نجارة", "أبواب"]):
            return "carpentry"
        return "general"

    def _get_grouped_items(self, cost_breakdown: Dict[str, Any]) -> OrderedDict:
        """
        Extract all items from cost breakdown grouped by category.
        Returns OrderedDict[category_key -> list of item dicts].
        """
        groups: OrderedDict = OrderedDict()

        # Materials items — group by detected trade
        if "materials" in cost_breakdown and "items" in cost_breakdown["materials"]:
            for item in cost_breakdown["materials"]["items"]:
                cat_key = self._detect_item_category(item)
                groups.setdefault(cat_key, []).append({
                    "description": item.get("description", item.get("name", "")),
                    "name": item.get("name", ""),
                    "quantity": item.get("quantity", 0),
                    "unit": item.get("unit", ""),
                    "unit_price": item.get("unit_price", 0),
                    "total": item.get("total", 0),
                })

        # Labor items
        if "labor" in cost_breakdown and "trades" in cost_breakdown["labor"]:
            for trade in cost_breakdown["labor"]["trades"]:
                groups.setdefault("labor", []).append({
                    "description": trade.get("description", f"{trade.get('trade', '').replace('_', ' ').title()}"),
                    "name": trade.get("trade", "").replace("_", " ").title(),
                    "quantity": trade.get("quantity", 0),
                    "unit": trade.get("unit", "Hour"),
                    "unit_price": trade.get("unit_price", 0),
                    "total": trade.get("total", 0),
                })

        # Permits and fees
        if "permits_and_fees" in cost_breakdown:
            if "items" in cost_breakdown["permits_and_fees"]:
                for item in cost_breakdown["permits_and_fees"]["items"]:
                    groups.setdefault("permits", []).append({
                        "description": item.get("description", item.get("name", "")),
                        "name": item.get("name", ""),
                        "quantity": item.get("quantity", 1),
                        "unit": item.get("unit", "Item"),
                        "unit_price": item.get("unit_cost", item.get("cost", 0)),
                        "total": item.get("cost", 0),
                    })
            elif cost_breakdown["permits_and_fees"].get("subtotal", 0) > 0:
                groups.setdefault("permits", []).append({
                    "description": "Permits and Regulatory Fees / التراخيص والرسوم التنظيمية",
                    "name": "Permits & Fees",
                    "quantity": 1,
                    "unit": "Lump Sum",
                    "unit_price": cost_breakdown["permits_and_fees"].get("subtotal", 0),
                    "total": cost_breakdown["permits_and_fees"].get("subtotal", 0),
                })

        return groups

    def _get_all_items(self, cost_breakdown: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Flat list of all items (backward compat for Materials/Labor sheets)."""
        all_items = []
        item_number = 1
        grouped = self._get_grouped_items(cost_breakdown)
        for items in grouped.values():
            for item in items:
                all_items.append({**item, "item_no": item_number, "category": "Materials"})
                item_number += 1

        # Contingency
        if "contingency" in cost_breakdown and cost_breakdown["contingency"].get("subtotal", 0) > 0:
            all_items.append({
                "item_no": item_number, "category": "Contingency",
                "description": f"Contingency ({cost_breakdown['contingency'].get('percentage', 0):.1f}%)",
                "name": "Contingency", "quantity": 1, "unit": "Lump Sum",
                "unit_price": cost_breakdown["contingency"]["subtotal"],
                "total": cost_breakdown["contingency"]["subtotal"],
            })
            item_number += 1

        # Markup
        if "markup" in cost_breakdown and cost_breakdown["markup"].get("subtotal", 0) > 0:
            all_items.append({
                "item_no": item_number, "category": "Markup",
                "description": f"Markup ({cost_breakdown['markup'].get('percentage', 0):.1f}%)",
                "name": "Markup", "quantity": 1, "unit": "Lump Sum",
                "unit_price": cost_breakdown["markup"]["subtotal"],
                "total": cost_breakdown["markup"]["subtotal"],
            })
            item_number += 1

        return all_items
    
    def generate_quotation_excel(self, quotation: Quotation, quotation_data: Optional[QuotationData]) -> BytesIO:
        """Generate Excel quotation with multiple sheets"""
        buffer = BytesIO()
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Detect language
        detected_lang = detect_language(quotation.project_description)
        is_arabic = detected_lang == "ar"
        
        # Normalize cost breakdown
        cost_breakdown = {}
        if quotation_data and quotation_data.cost_breakdown:
            total_cost = quotation_data.total_cost or 0.0
            cost_breakdown = self._normalize_cost_breakdown(quotation_data.cost_breakdown, total_cost)
        
        # Create main BOQ sheet first (most important)
        self._create_boq_sheet(wb, quotation, quotation_data, cost_breakdown, is_arabic)
        
        # Create summary sheet
        self._create_summary_sheet(wb, quotation, quotation_data, is_arabic)
        
        # Create detailed breakdown sheets
        if cost_breakdown:
            if "materials" in cost_breakdown and cost_breakdown["materials"].get("items"):
                self._create_materials_sheet(wb, cost_breakdown["materials"], is_arabic)
            
            if "labor" in cost_breakdown and cost_breakdown["labor"].get("trades"):
                self._create_labor_sheet(wb, cost_breakdown["labor"], is_arabic)
            
            self._create_breakdown_sheet(wb, cost_breakdown, is_arabic)
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _create_boq_sheet(self, wb: Workbook, quotation: Quotation, quotation_data: Optional[QuotationData],
                          cost_breakdown: Dict[str, Any], is_arabic: bool):
        """Create hierarchical BOQ sheet grouped by trade/category with bilingual headers."""
        ws = wb.create_sheet("BOQ - قائمة الكميات", 0)

        row = 1

        # Title section
        ws.merge_cells(f'A{row}:G{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = "BILL OF QUANTITIES / قائمة الكميات والأسعار"
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 2

        # Project information section
        ws[f'A{row}'] = "Quotation ID / رقم العرض:"
        ws[f'B{row}'] = quotation.id
        ws[f'A{row}'].font = self.bold_font
        row += 1

        ws[f'A{row}'] = "Date / التاريخ:"
        ws[f'B{row}'] = quotation.created_at.strftime('%Y-%m-%d')
        ws[f'A{row}'].font = self.bold_font
        row += 1

        ws[f'A{row}'] = "Project Description / وصف المشروع:"
        ws[f'B{row}'] = quotation.project_description
        ws[f'A{row}'].font = self.bold_font
        ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        row += 1

        if quotation.location:
            ws[f'A{row}'] = "Location / الموقع:"
            ws[f'B{row}'] = quotation.location
            ws[f'A{row}'].font = self.bold_font
            row += 1

        if quotation.project_type:
            ws[f'A{row}'] = "Project Type / نوع المشروع:"
            ws[f'B{row}'] = quotation.project_type.value.replace('_', ' ').title()
            ws[f'A{row}'].font = self.bold_font
            row += 1

        if quotation.timeline:
            ws[f'A{row}'] = "Timeline / الجدول الزمني:"
            ws[f'B{row}'] = quotation.timeline
            ws[f'A{row}'].font = self.bold_font
            row += 1

        row += 1

        # Bilingual BOQ Table Headers
        headers = [
            "Item No.\nرقم البند",
            "Description / الوصف",
            "Quantity\nالكمية",
            "Unit\nالوحدة",
            "Unit Price (EGP)\nسعر الوحدة",
            "Total (EGP)\nالإجمالي",
        ]

        header_row = row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
        ws.row_dimensions[header_row].height = 36
        row += 1

        # ── Hierarchical items by category ──────────────────────────
        grouped = self._get_grouped_items(cost_breakdown) if cost_breakdown else OrderedDict()

        category_num = 0
        item_data_rows: List[int] = []  # rows with =D*F formulas (for subtotal SUM)
        subtotal_rows: List[int] = []  # category subtotal rows (for grand total SUM)
        item_counter = 0  # for alternating rows

        for cat_key, items in grouped.items():
            category_num += 1
            names = _CATEGORY_NAMES_BILINGUAL.get(cat_key, {"en": cat_key.title(), "ar": cat_key})
            cat_label = f"{names['ar']} / {names['en']}"

            # ── Category header row (merged, bold, shaded) ──
            ws.merge_cells(f'A{row}:F{row}')
            cat_cell = ws.cell(row=row, column=1, value=f"{category_num}   {cat_label}")
            cat_cell.font = self.category_font
            cat_cell.fill = self.category_fill
            cat_cell.alignment = Alignment(horizontal='left', vertical='center')
            for c in range(1, 7):
                ws.cell(row=row, column=c).border = self.border
                ws.cell(row=row, column=c).fill = self.category_fill
            ws.row_dimensions[row].height = 28
            row += 1

            cat_start_row = row  # first data row in this category

            for sub_idx, item in enumerate(items, 1):
                item_counter += 1
                item_no = f"{category_num}.{sub_idx}"

                ws.cell(row=row, column=1, value=item_no).alignment = Alignment(horizontal='center', vertical='top')
                desc_cell = ws.cell(row=row, column=2, value=item.get("description", item.get("name", "")))
                desc_cell.alignment = Alignment(wrap_text=True, vertical='top')
                ws.cell(row=row, column=3, value=item.get("quantity", 0)).alignment = Alignment(horizontal='right')
                ws.cell(row=row, column=4, value=item.get("unit", "")).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=5, value=item.get("unit_price", 0)).number_format = '#,##0.00'
                ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
                # Formula: =C{row}*E{row}
                ws.cell(row=row, column=6).value = f"=C{row}*E{row}"
                ws.cell(row=row, column=6).number_format = '#,##0.00'
                ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')

                # Alternating row colors
                if item_counter % 2 == 0:
                    for c in range(1, 7):
                        ws.cell(row=row, column=c).fill = self.row_even_fill

                # Borders
                for c in range(1, 7):
                    ws.cell(row=row, column=c).border = self.border

                # Dynamic row height based on description length
                desc_len = len(str(item.get("description", "")))
                if desc_len > 200:
                    ws.row_dimensions[row].height = 60
                elif desc_len > 100:
                    ws.row_dimensions[row].height = 45
                else:
                    ws.row_dimensions[row].height = 30

                item_data_rows.append(row)
                row += 1

            # No items edge case
            if not items:
                continue

            cat_end_row = row - 1  # last data row in this category
            subtotal_rows.append(row)
            row += 1  # leave a blank row after each category

        # ── Subtotal ──
        subtotal_row = row
        ws.merge_cells(f'A{subtotal_row}:E{subtotal_row}')
        ws.cell(row=subtotal_row, column=1, value="Subtotal / المجموع الفرعي")
        ws.cell(row=subtotal_row, column=1).font = self.bold_font
        ws.cell(row=subtotal_row, column=1).alignment = Alignment(horizontal='right')
        # SUM of all item data rows
        if item_data_rows:
            sum_parts = ",".join(f"F{r}" for r in item_data_rows)
            ws.cell(row=subtotal_row, column=6).value = f"=SUM({sum_parts})"
        else:
            ws.cell(row=subtotal_row, column=6).value = 0
        ws.cell(row=subtotal_row, column=6).number_format = '#,##0.00'
        ws.cell(row=subtotal_row, column=6).font = self.bold_font
        ws.cell(row=subtotal_row, column=6).alignment = Alignment(horizontal='right')
        for c in range(1, 7):
            ws.cell(row=subtotal_row, column=c).border = self.border
        row += 1

        # ── Contingency ──
        contingency_row = None
        if "contingency" in cost_breakdown and cost_breakdown["contingency"].get("subtotal", 0) > 0:
            contingency_row = row
            pct = cost_breakdown["contingency"].get("percentage", 10)
            ws.merge_cells(f'A{row}:E{row}')
            ws.cell(row=row, column=1, value=f"Contingency / الطوارئ ({pct:.0f}%)")
            ws.cell(row=row, column=1).font = self.bold_font
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=6).value = f"=F{subtotal_row}*{pct/100}"
            ws.cell(row=row, column=6).number_format = '#,##0.00'
            ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
            for c in range(1, 7):
                ws.cell(row=row, column=c).border = self.border
            row += 1

        # ── Markup ──
        markup_row = None
        if "markup" in cost_breakdown and cost_breakdown["markup"].get("subtotal", 0) > 0:
            markup_row = row
            pct = cost_breakdown["markup"].get("percentage", 10)
            ws.merge_cells(f'A{row}:E{row}')
            ws.cell(row=row, column=1, value=f"Markup / هامش الربح ({pct:.0f}%)")
            ws.cell(row=row, column=1).font = self.bold_font
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=6).value = f"=F{subtotal_row}*{pct/100}"
            ws.cell(row=row, column=6).number_format = '#,##0.00'
            ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
            for c in range(1, 7):
                ws.cell(row=row, column=c).border = self.border
            row += 1

        # ── Grand Total ──
        total_row = row
        ws.merge_cells(f'A{total_row}:E{total_row}')
        ws.cell(row=total_row, column=1, value="GRAND TOTAL / الإجمالي الكلي")
        ws.cell(row=total_row, column=1).font = self.subtitle_font
        ws.cell(row=total_row, column=1).alignment = Alignment(horizontal='right')

        # Build SUM formula: subtotal + contingency + markup
        sum_refs = [f"F{subtotal_row}"]
        if contingency_row:
            sum_refs.append(f"F{contingency_row}")
        if markup_row:
            sum_refs.append(f"F{markup_row}")
        ws.cell(row=total_row, column=6).value = f"=SUM({','.join(sum_refs)})"
        ws.cell(row=total_row, column=6).number_format = '#,##0.00'
        ws.cell(row=total_row, column=6).font = self.subtitle_font
        ws.cell(row=total_row, column=6).alignment = Alignment(horizontal='right')

        total_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        for c in range(1, 7):
            cell = ws.cell(row=total_row, column=c)
            cell.border = self.thick_border
            cell.fill = total_fill

        # ── Column widths ──
        ws.column_dimensions['A'].width = 12   # Item No
        ws.column_dimensions['B'].width = 60   # Description
        ws.column_dimensions['C'].width = 12   # Quantity
        ws.column_dimensions['D'].width = 12   # Unit
        ws.column_dimensions['E'].width = 18   # Unit Price
        ws.column_dimensions['F'].width = 18   # Total
    
    def _create_summary_sheet(self, wb: Workbook, quotation: Quotation, quotation_data: Optional[QuotationData], is_arabic: bool):
        """Create comprehensive summary sheet"""
        ws = wb.create_sheet("Summary" if not is_arabic else "الملخص")
        
        row = 1
        
        # Title
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = "Construction Quotation Summary" if not is_arabic else "ملخص عرض السعر"
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 2
        
        # Project Information Section
        section_title = "Project Information" if not is_arabic else "معلومات المشروع"
        ws[f'A{row}'] = section_title
        ws[f'A{row}'].font = self.subtitle_font
        ws[f'A{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        # Quotation details
        ws[f'A{row}'] = "Quotation ID:" if not is_arabic else "رقم العرض:"
        ws[f'B{row}'] = quotation.id
        ws[f'A{row}'].font = self.bold_font
        row += 1
        
        ws[f'A{row}'] = "Date:" if not is_arabic else "التاريخ:"
        ws[f'B{row}'] = quotation.created_at.strftime('%Y-%m-%d')
        ws[f'A{row}'].font = self.bold_font
        row += 1
        
        ws[f'A{row}'] = "Description:" if not is_arabic else "الوصف:"
        ws[f'B{row}'] = quotation.project_description
        ws[f'A{row}'].font = self.bold_font
        ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        row += 1
        
        if quotation.location:
            ws[f'A{row}'] = "Location:" if not is_arabic else "الموقع:"
            ws[f'B{row}'] = quotation.location
            ws[f'A{row}'].font = self.bold_font
            row += 1
        
        if quotation.zip_code:
            ws[f'A{row}'] = "Zip Code:" if not is_arabic else "الرمز البريدي:"
            ws[f'B{row}'] = quotation.zip_code
            ws[f'A{row}'].font = self.bold_font
            row += 1
        
        if quotation.project_type:
            ws[f'A{row}'] = "Project Type:" if not is_arabic else "نوع المشروع:"
            ws[f'B{row}'] = quotation.project_type.value.replace('_', ' ').title()
            ws[f'A{row}'].font = self.bold_font
            row += 1
        
        if quotation.timeline:
            ws[f'A{row}'] = "Timeline:" if not is_arabic else "الجدول الزمني:"
            ws[f'B{row}'] = quotation.timeline
            ws[f'A{row}'].font = self.bold_font
            row += 1
        
        row += 1
        
        # Cost Summary Section
        if quotation_data and quotation_data.total_cost:
            cost_title = "Cost Summary" if not is_arabic else "ملخص التكلفة"
            ws[f'A{row}'] = cost_title
            ws[f'A{row}'].font = self.subtitle_font
            ws[f'A{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            
            ws[f'A{row}'] = "Total Cost:" if not is_arabic else "التكلفة الإجمالية:"
            ws[f'B{row}'] = f"{quotation_data.total_cost:,.2f} EGP"
            ws[f'A{row}'].font = self.bold_font
            ws[f'B{row}'].font = self.subtitle_font
            ws[f'B{row}'].number_format = '#,##0.00 "EGP"'
            row += 1
            
            if quotation_data.confidence_score is not None:
                ws[f'A{row}'] = "Confidence Score:" if not is_arabic else "نقاط الثقة:"
                ws[f'B{row}'] = f"{quotation_data.confidence_score:.1f}%"
                ws[f'A{row}'].font = self.bold_font
                row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
    
    def _create_materials_sheet(self, wb: Workbook, materials: Dict[str, Any], is_arabic: bool):
        """Create materials breakdown sheet"""
        ws = wb.create_sheet("Materials" if not is_arabic else "المواد")
        
        row = 1
        
        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = "Materials Breakdown" if not is_arabic else "تفاصيل المواد"
        title_cell.font = self.subtitle_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 2
        
        # Headers
        headers = [
            "Item No." if not is_arabic else "رقم البند",
            "Material" if not is_arabic else "المادة",
            "Category" if not is_arabic else "الفئة",
            "Quantity" if not is_arabic else "الكمية",
            "Unit" if not is_arabic else "الوحدة",
            "Unit Price (EGP)" if not is_arabic else "سعر الوحدة (جنيه)",
            "Total (EGP)" if not is_arabic else "الإجمالي (جنيه)"
        ]
        
        header_row = row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        row += 1
        
        # Data rows
        items = materials.get("items", [])
        for idx, item in enumerate(items, 1):
            ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=2, value=item.get("name", ""))
            ws.cell(row=row, column=3, value=item.get("category", ""))
            ws.cell(row=row, column=4, value=item.get("quantity", 0)).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=5, value=item.get("unit", "")).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=6, value=item.get("unit_cost", 0)).number_format = '#,##0.00'
            ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=7, value=item.get("cost", 0)).number_format = '#,##0.00'
            ws.cell(row=row, column=7).alignment = Alignment(horizontal='right')
            
            # Apply borders
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = self.border
            row += 1
        
        # Total row
        if items:
            total_row = row
            ws.merge_cells(f'A{total_row}:E{total_row}')
            ws.cell(row=total_row, column=1, value="Subtotal:" if not is_arabic else "المجموع الفرعي:")
            ws.cell(row=total_row, column=1).font = self.bold_font
            ws.cell(row=total_row, column=1).alignment = Alignment(horizontal='right')
            ws.cell(row=total_row, column=7, value=materials.get("subtotal", 0))
            ws.cell(row=total_row, column=7).number_format = '#,##0.00'
            ws.cell(row=total_row, column=7).font = self.bold_font
            ws.cell(row=total_row, column=7).alignment = Alignment(horizontal='right')
            
            for col in range(1, 8):
                ws.cell(row=total_row, column=col).border = self.border
        
        # Adjust column widths
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_labor_sheet(self, wb: Workbook, labor: Dict[str, Any], is_arabic: bool):
        """Create labor breakdown sheet"""
        ws = wb.create_sheet("Labor" if not is_arabic else "العمالة")
        
        row = 1
        
        # Title
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = "Labor Breakdown" if not is_arabic else "تفاصيل العمالة"
        title_cell.font = self.subtitle_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 2
        
        # Headers
        headers = [
            "Item No." if not is_arabic else "رقم البند",
            "Trade" if not is_arabic else "الحرفة",
            "Hours" if not is_arabic else "الساعات",
            "Rate (EGP/hr)" if not is_arabic else "المعدل (جنيه/ساعة)",
            "Cost (EGP)" if not is_arabic else "التكلفة (جنيه)"
        ]
        
        header_row = row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        row += 1
        
        # Data rows
        trades = labor.get("trades", [])
        for idx, trade in enumerate(trades, 1):
            ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=2, value=trade.get("trade", "").replace("_", " ").title())
            ws.cell(row=row, column=3, value=trade.get("hours", 0)).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=4, value=trade.get("rate", 0)).number_format = '#,##0.00'
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=5, value=trade.get("cost", 0)).number_format = '#,##0.00'
            ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
            
            # Apply borders
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self.border
            row += 1
        
        # Total row
        if trades:
            total_row = row
            ws.merge_cells(f'A{total_row}:D{total_row}')
            ws.cell(row=total_row, column=1, value="Subtotal:" if not is_arabic else "المجموع الفرعي:")
            ws.cell(row=total_row, column=1).font = self.bold_font
            ws.cell(row=total_row, column=1).alignment = Alignment(horizontal='right')
            ws.cell(row=total_row, column=5, value=labor.get("subtotal", 0))
            ws.cell(row=total_row, column=5).number_format = '#,##0.00'
            ws.cell(row=total_row, column=5).font = self.bold_font
            ws.cell(row=total_row, column=5).alignment = Alignment(horizontal='right')
            
            for col in range(1, 6):
                ws.cell(row=total_row, column=col).border = self.border
        
        # Adjust column widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _create_breakdown_sheet(self, wb: Workbook, cost_breakdown: Dict[str, Any], is_arabic: bool):
        """Create cost breakdown summary sheet"""
        ws = wb.create_sheet("Breakdown" if not is_arabic else "التفاصيل")
        
        row = 1
        
        # Title
        ws.merge_cells('A1:C1')
        title_cell = ws['A1']
        title_cell.value = "Cost Breakdown Summary" if not is_arabic else "ملخص تفاصيل التكلفة"
        title_cell.font = self.subtitle_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 2
        
        # Headers
        headers = [
            "Category" if not is_arabic else "الفئة",
            "Amount (EGP)" if not is_arabic else "المبلغ (جنيه)",
            "Percentage" if not is_arabic else "النسبة"
        ]
        
        header_row = row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        row += 1
        
        # Calculate total for percentage calculation
        total = sum([
            cost_breakdown.get("materials", {}).get("subtotal", 0),
            cost_breakdown.get("labor", {}).get("subtotal", 0),
            cost_breakdown.get("permits_and_fees", {}).get("subtotal", 0),
            cost_breakdown.get("contingency", {}).get("subtotal", 0),
            cost_breakdown.get("markup", {}).get("subtotal", 0)
        ])
        
        # Materials
        if "materials" in cost_breakdown and cost_breakdown["materials"].get("subtotal", 0) > 0:
            mat_label = "Materials" if not is_arabic else "المواد"
            ws.cell(row=row, column=1, value=mat_label)
            ws.cell(row=row, column=2, value=cost_breakdown["materials"].get("subtotal", 0))
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            percentage = (cost_breakdown["materials"].get("subtotal", 0) / total * 100) if total > 0 else 0
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%")
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = self.border
            row += 1
        
        # Labor
        if "labor" in cost_breakdown and cost_breakdown["labor"].get("subtotal", 0) > 0:
            lab_label = "Labor" if not is_arabic else "العمالة"
            ws.cell(row=row, column=1, value=lab_label)
            ws.cell(row=row, column=2, value=cost_breakdown["labor"].get("subtotal", 0))
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            percentage = (cost_breakdown["labor"].get("subtotal", 0) / total * 100) if total > 0 else 0
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%")
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = self.border
            row += 1
        
        # Permits
        if "permits_and_fees" in cost_breakdown and cost_breakdown["permits_and_fees"].get("subtotal", 0) > 0:
            perm_label = "Permits & Fees" if not is_arabic else "التراخيص والرسوم"
            ws.cell(row=row, column=1, value=perm_label)
            ws.cell(row=row, column=2, value=cost_breakdown["permits_and_fees"].get("subtotal", 0))
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            percentage = (cost_breakdown["permits_and_fees"].get("subtotal", 0) / total * 100) if total > 0 else 0
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%")
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = self.border
            row += 1
        
        # Contingency
        if "contingency" in cost_breakdown and cost_breakdown["contingency"].get("subtotal", 0) > 0:
            cont_label = "Contingency" if not is_arabic else "الطوارئ"
            ws.cell(row=row, column=1, value=cont_label)
            ws.cell(row=row, column=2, value=cost_breakdown["contingency"].get("subtotal", 0))
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            percentage = (cost_breakdown["contingency"].get("subtotal", 0) / total * 100) if total > 0 else 0
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%")
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = self.border
            row += 1
        
        # Markup
        if "markup" in cost_breakdown and cost_breakdown["markup"].get("subtotal", 0) > 0:
            mark_label = "Markup" if not is_arabic else "الهامش"
            ws.cell(row=row, column=1, value=mark_label)
            ws.cell(row=row, column=2, value=cost_breakdown["markup"].get("subtotal", 0))
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            percentage = (cost_breakdown["markup"].get("subtotal", 0) / total * 100) if total > 0 else 0
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%")
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = self.border
            row += 1
        
        # Total
        total_label = "TOTAL" if not is_arabic else "الإجمالي"
        ws.cell(row=row, column=1, value=total_label).font = self.subtitle_font
        ws.cell(row=row, column=2, value=total).number_format = '#,##0.00'
        ws.cell(row=row, column=2).font = self.subtitle_font
        ws.cell(row=row, column=3, value="100.0%").font = self.subtitle_font
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = self.thick_border
            ws.cell(row=row, column=col).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
